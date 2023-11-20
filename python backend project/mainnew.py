import csv
import openpyxl
import os
class Course:
    courseno=""
    file=""
    def __init__(self,courseno):
        self.__courseno=courseno
        self.file=self.__courseno+'.csv'
    
    def get_all_sections(self):
        try:
            with open(self.file,'r') as csvfile:
                data=csv.reader(csvfile)
                for i in data:
                    if i==[]:
                        pass
                    else:
                        print(i[0])
        except:
            print("No Courses Found with the input Parameters")
    
    def populate_section(self,newsec):
        path=self.file
        if os.path.exists(path):
            with open(path,'a') as csvfile:
                csvwriter=csv.writer(csvfile)
                data=[]
                room=input("Enter room no. ")
                day=[]
                times=int(input("No. of days it occurs: "))
                for i in range(0,times):
                    value=input("Enter the day(For Monday-M Tuesday-T Wednesday-W Thursday-Th Friday-F,Saturday-S) ")
                    day.append(value)
                hours=input("Enter the hours in list format.For example: [1] refers to first hour i.e. 8-8:50 and so on ")
                type=input('Enter its type (LAB TUT or LEC) ')
                data=[newsec,room,day,hours,type]
                csvwriter.writerow(data)
                print("\t SUCCESSFULLY CREATED!!! \t")
        else:
            print("No course found")
        
    #def __str__(self)



                
class Timetable:
    tt={}
    M=[]
    T=[]
    W=[]
    Th=[]
    F=[]
    S=[]
    flag={}
    def __init__(self):
        self.flag={}
        self.tt={}
        self.tt['M']={}
        self.tt['T']={}
        self.tt['W']={}
        self.tt['Th']={}
        self.tt['F']={}
        self.tt['S']={}
        self.flag['M']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['T']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['W']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['Th']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['F']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['S']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}


    def enroll_subject(self,courseno,sec):
        count=0
        file=courseno+'.csv'
        if os.path.exists(file):
            with open(file,'r') as csvfile:
                data=csv.reader(csvfile)
                count=0
                for i in data:
                    if i==[]:
                        pass
                    else:
                        if i[0]==sec:
                            count+=1
                            for j in eval(i[2]):
                                for k in eval(i[3]):
                                    if j=='M':
                                        self.tt['M'][k]=courseno + i[4]
                                        self.flag['M'][k]+=1
                                    elif j=='T':
                                        self.tt['T'][k]=courseno + i[4]
                                        self.flag['T'][k]+=1
                                    elif j=='W':
                                        self.tt['W'][k]=courseno + i[4]
                                        self.flag['W'][k]+=1
                                    elif j=='Th':
                                        self.tt['Th'][k]=courseno + i[4]
                                        self.flag['Th'][k]+=1
                                    elif j=='F':
                                        self.tt['F'][k]=courseno + i[4]
                                        self.flag['F'][k]+=1
                                    elif j=='S':
                                        self.tt['S'][k]=courseno + i[4]
                                        self.flag['S'][k]+=1
                                    else:
                                        pass
                if count==0:
                    print("Section entered is unavailable")
                else:
                    print("\t SUCCESSFULLY ADDED TO CART!!! \t")

    def check_clashes(self):
        count=0
        for i in self.flag:
            for j in self.flag[i]:
                if self.flag[i][j]>1:
                    print("clash found in ",i)
                    count+=1
        if count==0:
            print("no clashes")
    
    def export_to_csv(self):
        with open('timetable.csv','w') as file:
            writer=csv.writer(file)
            writer.writerow(['Day','8:00-8:50','9:00-9:50','10:00-10:50','11:00-11:50','12:00-12:50','1:00-1:50','2:00-2:50','3:00-3:50','4:00-4:50'])
            for i in self.tt:
                blank=['','','','','','','','','','']
                blank[0]=i
                for j in self.tt[i]:
                    blank[int(j)]=self.tt[i][j]
                writer.writerow(blank)
            print("\t SUCCESSFULLY EXPORTED TO CSV!!! \t")
    
    def clear(self):
        with open('timetable.csv','w') as file:
            writer=csv.writer(file)
            writer.writerow(['Day','8:00-8:50','9:00-9:50','10:00-10:50','11:00-11:50','12:00-12:50','1:00-1:50','2:00-2:50','3:00-3:50','4:00-4:50'])
        self.flag={}
        self.tt={}
        self.tt['M']={}
        self.tt['T']={}
        self.tt['W']={}
        self.tt['Th']={}
        self.tt['F']={}
        self.tt['S']={}
        self.flag['M']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['T']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['W']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['Th']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['F']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        self.flag['S']={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0}
        print("\t SUCCESSFULLY CLEARED!!! \t")


def populate_course(subject):
    path='allcourses.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    flag=0
 
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
 
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        if cell_obj.value==subject:
            flag+=1
            file=subject+'.csv'

            if os.path.exists(file):
                print("Course already exists")
            else:
                with open(file,'w') as csvfile:
                    csvfile.close()
                    print("\t SUCCESSFULLY CREATED!!! \t")
                    break
    if flag==0:
            print("Course not in excel sheet")             

            
                    
#CLI
while True:
    print("---------\t WELCOME TO TIMETABLE MANAGEMENT SYSTEM \t---------")
    print("Enter 1 to go to course section")
    print("Enter 2 to go to Timetable section to create timetable")
    print("Enter 3 to add a Course")
    print("Enter 0 to exit")
    print("---------\t -------------------------------------- \t---------")
    ui=input("Enter your choice--> \t")
    if ui=='0':
        print("Thank you!!")
        break
    elif ui=='1':
        courseno=input("Enter course no")
        a=Course(courseno)
        while True:
            print("Enter 1 to check all sections available")
            print("Enter 2 to add a section to the course")
            print("Enter 0 to exit to main menu")
            ui1=input("Enter your choice ")
            if ui1=='0':
                break
            elif ui1=='1':
                a.get_all_sections()
            elif ui1=='2':
                newsec=input("Enter new section's name ")
                a.populate_section(newsec)
            else:
                print("Invalid Input")
    elif ui=='2':
        a=Timetable()
        while True:
            print("Enter 1 to add a new course")
            print("Enter 2 to check for clashes")
            print("Enter 3 to add the timetable to csv.Note:Do this after confirming that there are no clashes.If there are clashes please clear the timetable")
            print("Enter 4 to clear timetable and cart")
            print("Enter 0 to exit")
            ui2=input("Enter your choice ")
            if ui2=='1':
                courseno=input("Enter course no. ")
                sec=input("Enter its Desired Section: ")

                a.enroll_subject(courseno,sec)
            
            elif ui2=='2':
                a.check_clashes()
            
            elif ui2=='3':
                a.export_to_csv()
            elif ui2=='4':
                a.clear()
            elif ui2=='0':
                break
    elif ui=='3':
        courseno=input("Enter course no. ")
        populate_course(courseno)
    
    else:
        print("Invalid Input")
        








        

    





                

    

    
